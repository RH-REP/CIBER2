from cmath import sqrt
import openpyxl 
import glob
import sys
import re
import matplotlib.pyplot as plt
# import matplotlib
import gc
from . import fits_reader
from collections import defaultdict
import numpy as np
import datetime 
import os
from memory_profiler import profile
from matplotlib.offsetbox import (TextArea,  AnnotationBbox)
import matplotlib.patches as patches

ch_dict = {"armL":"ch1","armM":"ch2","armS":"ch3",}
# pixelToWavelenght = {"armL:"}
def convert_pixel_to_lvf(pixel,arm_name):
    a = {
        "armL":[-4608.3,-2.124],
        "armM":[-3734.8,-2.37],
        "armS":[1225.0,3.23]
    }
    wavelength = (pixel+a[arm_name][0])/a[arm_name][1]
    return wavelength

def getColumnCoordinate(v):
    return re.split('[0-9]',v.coordinate)[0]

class FitsFolderClass():
    def __init__(self,):
        self._dataSet = []
        self._bg_dataSet = []
        return
    def write_data(self,data,bg_data):
        self._dataSet.append(data)
        self._bg_dataSet.append(bg_data)
        return
    def get_data_numbers(self,):
        return self._dataSet
    def get_bg_data_numbers(self,):
        return self._bg_dataSet

def make_reduction_shelf(experiment_directory,):
    return  Reduction_Shelf(experiment_directory)
    

def make_ramp_shelf(experiment_directory,):
    return Ramp_Shelf(experiment_directory)

def make_img_shelf(name,imgs,):
    return Image_Shelf(name,imgs)


# def make_line_shelf(name,x,ys,):
#     return Line_Shelf(name,x,ys)


class Reduction_Shelf():
    def __init__(self,experiment_directory):
        self._fitDatabook = defaultdict(FitsFolderClass)
        self._experiment_directory = experiment_directory
        self._line_up_books_by_itself()
        return

    def take_book(self,book_name):
        return self._fitDatabook[book_name]

    def get_all_books(self):
        ds = []
        for d,_ in self._fitDatabook.items():
            ds.append(d)
        return ds
    
    def get_eps_files(self,book_name,arm_name):
        eps_files = glob.glob(self._experiment_directory + book_name + "/" + ch_dict[arm_name]+"/*[0-9].f*ts")
        eps_files = sorted(eps_files,)
        eps_bg_files = glob.glob(self._experiment_directory + book_name + "/" + ch_dict[arm_name]+"/*bg.f*ts")
        eps_bg_files = sorted(eps_bg_files,)
        
        return eps_files, eps_bg_files

    def _read_imgs(self,files):
        imgs = np.full((len(files),2048,2048),np.nan)
        for i,f in enumerate(files):
            fits,_ = fits_reader.loadFits(f)
            imgs[i] = fits[0]
            
        return imgs
    
    def get_eps_img(self,book_name,arm_name,):
        """
        return img only
        """
        files ,bg_files=self.get_eps_files(book_name,arm_name)
        # print(files)
        imgs = self._read_imgs(files)
        img = np.nanmean(imgs,axis=0)

        if not(len(bg_files) ==0):
            bg_imgs = self._read_imgs(bg_files)
            bg_img = np.nanmean(bg_imgs,axis=0)

            img = img - bg_img
        else:
            print("bg is empty")
        return img


    def get_eps_imgs(self,book_name,arm_name,):
        """
        return img only
        """
        files ,bg_files=self.get_eps_files(book_name,arm_name)
        # print(files)
        imgs = self._read_imgs(files)
        return imgs

    def get_eps_imgs_sub_background(self,book_name,arm_name,):
        """
        return img only
        """
        files ,bg_files=self.get_eps_files(book_name,arm_name)
        # print(files)
        imgs = self._read_imgs(files)
        # img = np.nanmean(imgs,axis=0)

        if not(len(bg_files) ==0):
            bg_imgs = self._read_imgs(bg_files)
            # bg_img = np.nanmean(bg_imgs,axis=0)

            # img = img - bg_img
            imgs = imgs - bg_imgs

        else:
            print("bg is empty")
        return imgs



    # def get_lvf_img(self,book_name,arm_name,):
    #     img = self.get_eps_img(book_name,arm_name,)
    #     img_at_lvf = img[10:150,]
    #     return img_at_lvf 

    def get_lvf_imgs(self,book_name,arm_name,):
        imgs = self.get_eps_imgs_sub_background(book_name,arm_name,)
        img_at_lvf_s = imgs[:,10:150]
        return img_at_lvf_s

    # def show_eps_image(self,book_name,arm_name):
    #     img = self.get_eps_img(book_name,arm_name)
    #     fig,ax = plt.subplots()
    #     ax.imshow(img)
    #     return fig,ax

    def _line_up_books_by_itself(self,):
        folders = glob.glob(self._experiment_directory+"*")
        for f in folders:
            name = f.split("/")[-1]
            self.take_book(name)
        return 
    
    # def showLVFimage(self,book_name,arm_name,):
    #     img = self.get_eps_img(book_name,arm_name,)
    #     img_at_lvf = img[10:150,]
    #     fig, ax = plt.subplots()
    #     im = ax.imshow(img_at_lvf,cmap = 'jet')
    #     fig.colorbar(im,ax = ax,shrink=0.2)
    #     fig.set_size_inches(8,8)
    #     return 

    def getLowPixelVSIntensity(self,imgs,startPix=10,endPix=150):
        """
        default is LVF area
        """
        cropped_image_averages = []
        # print(imgs)
        for i,img in enumerate(imgs):
            cropped_image = img[startPix:endPix,]
            pixel = np.arange(len(img[0]))        
            cropped_image_average = np.nanmean(cropped_image,axis = 0)
            cropped_image_averages.append(cropped_image_average)
        return pixel,np.array(cropped_image_averages)

    

    def getLVFxy(self,book_name,arm_name,):
        imgs = self.get_eps_imgs_sub_background(book_name,arm_name)
        x_lvf_pix,y_lvf_s = self.getLowPixelVSIntensity(imgs)
        y_lvf = np.nanmean(y_lvf_s,axis = 0)
        x_lvf = convert_pixel_to_lvf(x_lvf_pix,arm_name)
        return x_lvf,y_lvf

    def getLVF_with_sterr(self,book_name,arm_name,):
        imgs = self.get_eps_imgs_sub_background(book_name,arm_name)
        x_lvf_pix,y_lvf_s = self.getLowPixelVSIntensity(imgs)
        y_lvf = np.nanmean(y_lvf_s,axis = 0)
        y_sterr = np.std(y_lvf_s,axis = 0)/np.sqrt(len(y_lvf_s))
        # print(y_lvf.T[1000])
        # print(len(y_lvf_s))

        x_lvf = convert_pixel_to_lvf(x_lvf_pix,arm_name)

        return x_lvf,y_lvf,y_sterr





    # def getLVFxy_for_each(self,book_name,arm_name,):
    #     imgs = self.get_eps_imgs(book_name,arm_name,)
    #     img_at_lvf_s = imgs[:,10:150,]
    #     pixel = np.arange(len(imgs[0][0]))
    #     x_lvf = convert_pixel_to_lvf(pixel,arm_name)
    #     return x_lvf,img_at_lvf_s

class Image_Shelf():

    def __init__(self,name,imgs):
        self.imgs = imgs
        self.names = name

        return
        
    def convert_volt_to_eps(self):
        for i,img in enumerate(self.imgs):
            self.imgs[i] = img * (4*10**(-6))

    def star_clip(self,threshould):
        for i,img in enumerate(self.imgs):

            img = np.where(img < threshould, img , np.nan)
            img = np.where(img > -threshould, img , np.nan)
            self.imgs[i]  = img

    def substract_bg(self):
        for i,img in enumerate(self.imgs):
            img = img - np.nanmean(img[:,10:100])
            self.imgs[i]  = img


    def substract_line_noise(self):
        for i,img in enumerate(self.imgs):
            line_noise = np.nanmean(img[:,0:4],axis=1) 
            # line_noise = np.nanmean(img[:,0:4]+img[:,-5:-1],axis=1) 
            v = np.array([1/4, 1/4, 1/4, 1/4, 1/4])
            line_noise_smooth = np.convolve(line_noise, v, mode ="valid")
            a = line_noise_smooth[0]
            b = line_noise_smooth[-1]
            line_noise_smooth =  np.insert(line_noise_smooth,0,([a,a]))
            line_noise_smooth =  np.insert(line_noise_smooth,-1,([b,b]))

            img = img - line_noise_smooth
            self.imgs[i]  = img

    def getLowPixelVSIntensity(self,startPix,endPix):
        cropped_image_averages = []
        for i,img in enumerate(self.imgs):
            cropped_image = img[startPix:endPix,]
            pixel = np.arange(len(img[0]))        
            cropped_image_average = np.nanmean(cropped_image,axis = 0)
            cropped_image_averages.append(cropped_image_average)
        return pixel,np.array(cropped_image_averages)


    def cut_out_square(self,x_startPix,x_endPix,y_startPix,y_endPix):
        cropped_images = []
        for i,img in enumerate(self.imgs):
            cropped_images.append(img[x_startPix:x_endPix,y_startPix:y_endPix])
        return np.array(cropped_images)

# class Line_Shelf():

#     def __init__(self,name,x,ys):
#         self.names = name
#         self.x = x
#         self.ys = ys
#         return

#     def getAverages(self,startPix,endPix):

#         pixel,img_averages = self.getLowPixelVSIntensity(startPix,endPix)

class Ramp_Shelf():
    def __init__(self,experiment_directory):
        self._fitDatabook = defaultdict(FitsFolderClass)
        self._experiment_directory = experiment_directory
        self.minus_reset_frame = True
        self.cut_area = None
        self.cut_area_L = fits_reader.AreaSizeClass(400,500,1400,1500)
        self.cut_area_R = fits_reader.AreaSizeClass(1400,1500,1400,1500)
        self.is_linear_fit  = True
        self.isMask = True
        return


    def take_book(self,book_name):
        return self._fitDatabook[book_name]
        
    def show_all_books(self):
        ds = []
        for d,_ in self._fitDatabook.items():
            ds.append(d)
        print(ds)
        return ds

    def show_numbers_in_book(self,book_name,):
        numbers = self.take_book(book_name).get_data_numbers()
        print("len is ", len(numbers))
        for number in numbers:
            print(number)
        return 

    def _get_folder_name(self,arm_name,number):
        nameA = self._experiment_directory +"/frames/" + ch_dict[arm_name] + "/*"
        if number == None:
                return []
        folder_name = glob.glob(nameA + str(number))
        if folder_name == []:
            folder_name = glob.glob(nameA + str(number+1))
        if folder_name == []:
            folder_name = glob.glob(nameA + str(number-1))
        return folder_name

    def _get_working_directory(self,book_name,arm_name,isBg = False):
        folders = []
        b = self.take_book(book_name)
        numbers = b.get_data_numbers() if isBg == False else b.get_bg_data_numbers()

        for number in numbers:
            folder_name_list = self._get_folder_name(arm_name,number)
            if not len(folder_name_list) == 0:
                folders.append(folder_name_list[0])
            else:
                print("no such directory :", number)

        return folders
    

    def make_ramp_graph(self,book_name,arm_name,isBg = False,):
        if self.cut_area == None:
            print("undefined cut area")
            return None, None
        fig, ax  = plt.subplots()
        working_directory = self._get_working_directory(book_name,arm_name,isBg = isBg)
        self.plot_ramp_graph(working_directory,arm_name,ax)
        # for f in working_directory:
        #     label = f.split("_")[-1]
        #     times, average_counts_of_area,slope =self.calc_ramp_data(f,arm_name)
        #     if times == []:
        #         continue
        #     lineobj = ax.plot(times, average_counts_of_area,label = label, marker='o', lw=0)
        #     if self.is_linear_fit:
        #         ax.plot(np.append(0,times),np.average(slope)*np.append(0,times),c = lineobj[0].get_c() , linestyle='dashed')
                
        # ax.set_ylabel('electrons')
        # ax.set_xlabel('time[s]')
        # title_name = book_name + "-"+ arm_name if isBg == False else book_name + "-"+ arm_name + "-BG"
        # ax.set_title(title_name)
        # ax.set_xlim(0,)
        # ax.set_ylim(0,)
        title_name = book_name + "-"+ arm_name if isBg == False else book_name + "-"+ arm_name + "-BG"
        ax.set_title(title_name)

        if len(working_directory) < 10:
            ax.legend()
        return fig, ax
    def plot_ramp_graph(self,working_directory,arm_name,ax):
        # print(working_directory)
        for i, f in enumerate(working_directory):
            label = f.split("_")[-1]
            times, average_counts_of_area,average_counts_of_area_std,slope =self.calc_ramp_data(f,arm_name)
            if times == []:
                continue
            lineobj = ax.plot(times, average_counts_of_area,label = label,ms=1, marker='o', lw=0)
            ax.errorbar(times, average_counts_of_area, yerr=average_counts_of_area_std,c = lineobj[0].get_c())
            if self.is_linear_fit:
                ax.plot(np.append(0,times),np.average(slope)*np.append(0,times),lw = 0.3 ,c = lineobj[0].get_c(), linestyle='dashed')
                if i == 0:
                    ax.text(times[-1],0 , 'eps = '+ str(round(np.average(slope),1)),
                    horizontalalignment='center',
                    verticalalignment='bottom',
                    multialignment='center')

        ax.set_ylabel('electrons')
        ax.set_xlabel('time[s]')
        ax.set_xlim(0,)
        ax.set_ylim(0,)
        if len(working_directory) < 10:
            ax.legend()
    

    def make_flight_sequence_graph(self,book_name,arm_name,isBg = False,):
        fig, ax  = plt.subplots()
        working_directory = self._get_working_directory(book_name,arm_name,isBg=isBg)
        times_np = np.array([])
        last_time = fits_reader.convertWorkingdirectoryToSecond(working_directory)
        print("start time ,",last_time)
        average_counts_of_area_np = np.array([])

        for f in working_directory:
            # times, average_counts_of_area, _ = self.calc_ramp_data(f,arm_name,)
            times, average_counts_of_area, = self.calc_ramp_data(f,arm_name,)
            times_np = np.append(times_np,times+last_time)
            last_time = times_np[-1]
            average_counts_of_area_np = np.append(average_counts_of_area_np,average_counts_of_area)

        ax.plot(times_np, average_counts_of_area_np, marker='o', lw=0)
        names = ["bootes"," bootes dither","Elat17"," swire","swire dither"," NEP","NEP dither"]
        cl = ['#e0ffff','#90ee90']
        xt = [105,170,227,262,309,367,415,715]
        print(np.nanmax(average_counts_of_area_np))
        for i in range(7):
            ax.text(xt[i], 0, " "*50+ names[i],rotation =90 )
            ax.axvspan(xt[i], xt[i+1], facecolor=cl[i%2])

        ax.set_ylabel('electrons')
        ax.set_xlabel('time[s]')
        title_name = book_name+ "-"+ arm_name 
        ax.set_title(title_name)
        ax.set_xlim(0,)
        ax.set_ylim(0,)
        return fig, ax


    # @profile
    def make_ramp_map(self,book_name,arm_name,isBg = False):
        fig, (ax,ax2)  = plt.subplots(1,2,figsize=(16, 8))
        ax.set_title(book_name + "-" + arm_name)
        ax2.set_title(book_name + "-" + arm_name +  "-y err")
        working_directory = self._get_working_directory(book_name,arm_name,isBg = isBg)
        for f in working_directory:
            self.plot_eps(arm_name,f,fig,ax,ax2)
            break
        return fig,(ax,ax2)
    

    def plot_rectangle(self,cut_area,ax,label):
        xs = cut_area.xs
        xe = cut_area.xe
        ys = cut_area.ys
        ye = cut_area.ye
        r = patches.Rectangle(xy=(xs, ys), width=xe-xs, height=ys-ye, ec='#000000', fill=False)
        ax.add_patch(r)
        offsetbox = TextArea(label)
        ab = AnnotationBbox(offsetbox, (xs+100, ys),
                            xybox=(-5, 30),
                            xycoords='data',
                            boxcoords="offset points",
                            )
        ax.add_artist(ab)
        return 
    
    def plot_eps(self,arm_name,f,fig,ax,ax2):

        times, data_image_minus_reset = self._get_data(arm_name,f)  
        if times == []:
            print("no files ")
            return 
        a,b,steyx,_ = self._fit_linear(times,data_image_minus_reset)
        im = ax.imshow(a,vmin = 0, vmax = np.nanmean(a)*2.2,cmap = "jet")
        cbar = fig.colorbar(im,ax = ax,shrink=0.8)
        cbar.set_label("eps")
        print("average std % :",np.nanmean(steyx/a))
        im2 = ax2.imshow(steyx/a*100,vmin = 0,vmax = 50,cmap = "jet")
        cbar2 = fig.colorbar(im2,ax = ax2,shrink=0.8)
        cbar2.set_label("std %")
        return a,b,steyx

    def show_LR_cut_area(self,book_name,arm_name,isBg = False):
        fig,axes = plt.subplots(2,2,sharex="row",sharey="row")
        # fig, (ax,ax2) =  self.make_ramp_map(book_name,arm_name,isBg )
        ax00 = axes[0][0]
        ax01 = axes[0][1]
        ax00.set_title(book_name + "-" + arm_name)
        ax01.set_title(book_name + "-" + arm_name +  "-y err")

        working_directory = self._get_working_directory(book_name,arm_name,isBg = isBg)
        f = working_directory[1]
        self.plot_eps(arm_name,f,fig,ax00,ax01)
        self.plot_rectangle(self.cut_area_L,ax00,"L")
        self.plot_rectangle(self.cut_area_R,ax00,"R")

        ax10 = axes[1][0]
        ax11 = axes[1][1]
        working_directory = self._get_working_directory(book_name,arm_name,isBg = isBg)
        cut_angle_save = self.cut_area
        self.cut_area = self.cut_area_R
        self.plot_ramp_graph(working_directory,arm_name,ax11)
        self.cut_area = self.cut_area_L
        self.plot_ramp_graph(working_directory,arm_name,ax10)
        self.cut_area = cut_angle_save
        
        title_name = book_name + "-"+ arm_name if isBg == False else book_name + "-"+ arm_name + "-BG"
        ax10.set_title(title_name + "-L")
        ax11.set_title(title_name + "-R")
        ax11.set_ylabel("")

        return fig, axes
        
    def save_linear_reduction(self,experiment_name,book_name,arm_name,isBg = False,isShow=False):
        home_directory = os.environ['HOME']
        season_name = "/WSMR2021May"
        reduction_directory = home_directory + season_name + "/reduction/" + str(datetime.date.today()) + experiment_name + "/" + book_name.replace(" ","_")+ "/" + ch_dict[arm_name] 
        working_directory = self._get_working_directory(book_name,arm_name,isBg = isBg,)
        os.makedirs(reduction_directory,exist_ok=True)
        # print(working_directory)
        for f in working_directory:
            number = f.split("_")[-1]
            print(number)
            times, data_image_minus_reset = self._get_data(arm_name,f)  
            if times == []:
                continue
            gc.collect()
            ramp_image,b,std_image,pixel_image =  self._fit_linear(times,data_image_minus_reset)
            if isShow:
                fig,ax = plt.subplots()
                ax.imshow(ramp_image)
            savename = reduction_directory  +"/" + number 
            if isBg:
                savename += "_bg"
            savename += ".fts"
            fits_reader.saveFits(ramp_image,savename,std_image,pixel_image)
        return

    def _get_data(self,arm_name,folder_name,):
        ch = ch_dict[arm_name]
        times, data_image= fits_reader.getTimeAndImage(folder_name)
        if times == [] or len(times) < 3:
            print("not enough files in", folder_name)
            return [],[]


        if self.isMask:
            mask,_ = fits_reader.loadMask(ch)
            data_image = data_image * mask
            if self.minus_reset_frame:
                reset_image = self._getResetImage(ch)
                data_image = data_image - reset_image
        return times, data_image


    def calc_ramp_data(self,folder_name,arm_name,):
        cut_area = self.cut_area
        times, data_image_minus_reset = self._get_data(arm_name,folder_name)        
        if times == []:
            return [],[],[]
        slope = np.array([])
        if self.is_linear_fit:
            slope,intercept =  self._fit_linear_in_area(times,data_image_minus_reset)
            print("slope is ",np.average(slope))
            time_shift = np.nanmean(intercept)/np.nanmean(slope)
            times = times + time_shift

        data_length = len(data_image_minus_reset)
        area_average_set = np.arange(data_length)
        area_average_set_error = np.arange(data_length)
        for i,data in enumerate(data_image_minus_reset):
           area_average_set[i],area_average_set_error[i] = np.nanmean(data[cut_area.ys:cut_area.ye,cut_area.xs:cut_area.xe]),np.nanstd(data[cut_area.ys:cut_area.ye,cut_area.xs:cut_area.xe])
        return times,area_average_set,#area_average_set_error,slope


# """
# 本整理2種類
# """


    def line_up_by_folders(self,book_name,folders,):
        """
        処理するフォルダーをあらかじめ決定し、名前をつけて整理する
        """
        ramp_book = self.take_book(book_name)
        for f in folders:
            ramp_book.write_data(f.split("_")[-1],None)
        return 
    
    def line_up_by_excel(self):

        """
        デフォルトの整理手段
        """
        filename = glob.glob(self._experiment_directory + "/*.xlsx")[-1]
        wb = openpyxl.load_workbook(filename)
        ws = wb.worksheets[0]
        first_row = ws[1]
        configuration_index = None
        d1_index = None
        d0_index = None
        for _, v in enumerate(first_row):
            if v.value == "Configuration":
                configuration_index = getColumnCoordinate(v)
            elif v.value == "Scan Number":
                d1_index = getColumnCoordinate(v)
            elif v.value == "BG Number":
                d0_index = getColumnCoordinate(v)
        if not(configuration_index and d1_index and d0_index):
            sys.exit("scan number error")
        configurations = ws[configuration_index][1:]
        data_set = ws[d1_index][1:]
        bg_data_set = ws[d0_index][1:]

        for (cf,dt_number,bg_number)in zip(configurations,data_set,bg_data_set):
            ramp_book = self.take_book(cf.value)
            ramp_book.write_data(dt_number.value,bg_number.value)
        return 

    def _getResetImage(self,ch):
        reset_fits_file = "./lib/reset_image/" + ch + "_reset.fits"
        fits,_ = fits_reader.loadFits(reset_fits_file)
        reset_image = fits[0]
        return reset_image
        
    def _fit_linear_in_area(self,x,y_np):
        datalen = len(x)
        cut_area = self.cut_area
        size = 100 
        y_np_reduced = y_np[:,cut_area.ys:cut_area.ye,cut_area.xs:cut_area.xe]
        y_d1shape = np.reshape(y_np_reduced,(datalen,size*size))
        gc.collect()
        linear_model,cov_params=np.polyfit(x,y_d1shape,1,cov=True)
        [a,b] = np.reshape(linear_model,(2,size,size))
        return a,b, 

    def _fit_linear(self,x,y_np):
        datalen = len(x)
        datalen = min(len(x),30)
        x = x[:datalen]
        y_np = y_np[:datalen]
        size = 2048
        y_len_vs_d1 = np.reshape(y_np,(datalen,size*size))
        linear_model,cov_params = np.polyfit(x,y_len_vs_d1,1,cov=True)
        # print("x:",x.shape)
        # print("y_len_vs_d1:",y_len_vs_d1.shape)
        # print("cov_params:",cov_params)
        std_image_array = np.reshape(cov_params[0][0],(size,size))
        [a,b] = np.reshape(linear_model,(2,size,size))
        pixel_numbers = np.full((size,size),datalen,dtype=np.int32)
        return a,b, std_image_array,pixel_numbers

        #""""Do NOT deleate"""""""
        # fig2,ax2 = plt.subplots()
        # fig3,ax3 = plt.subplots()
        # ax2.imshow(a)
        # ax3.imshow(b)
        #""""""""""""""""""""""""


