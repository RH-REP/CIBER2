import astropy.io.fits as fits
import numpy as np
import openpyxl 
import sys
import re
from collections import defaultdict

class AreaSizeClass:
    def __init__(self,xs,xe,ys,ye):
        self.xs = xs
        self.xe = xe
        self.ys = ys
        self.ye = ye

def loadFits(fileName):
    hdulist=fits.open(fileName)
    hdu=hdulist[0]
    data=np.array(hdu.data)
    header=hdu.header
    return data, header

def getFrameNumber(filename):
    return int(filename.split('/FRM')[-1].split('_PIX')[0])

def convetNumberToSecond(number):
    return number * 1.5

def getTimeAndImage(files, reset_fits):
    reset_image,_ = loadFits(reset_fits)
    reset_number = getFrameNumber(reset_fits)
    file_length = len(files)

    times = np.full(file_length,np.nan)
    data_image_minus_reset = np.full((file_length,2048,2048,),np.nan)

    # print(data_image_minus_reset)
    for i,f in enumerate(files):
        frame_number = getFrameNumber(f)
        data_image,_ = loadFits(f)
        times[i]= convetNumberToSecond(frame_number-reset_number)
        data_image_minus_reset[i] = data_image-reset_image
    return times, data_image_minus_reset

def getAverage(data,cutArea):
    data_length = len(data)
    area_average = np.arange(data_length)
    area_average_error = np.arange(data_length)
    for i,d in enumerate(data):
        area_average[i]= np.average(d[cutArea.ys:cutArea.ye,cutArea.xs:cutArea.xe])/(4*10**(-6))
        area_average_error[i]= np.std(d[cutArea.ys:cutArea.ye,cutArea.xs:cutArea.xe])/((4*10**(-6))*100)
    return area_average,area_average_error

def getColumnCoordinate(v):
    return re.split('[0-9]',v.coordinate)[0]

class FitsDataClass():
    # def __init__(self,dataSet,bg_dataSet):
    #     self._dataSet = dataSet
    #     self._bg_dataSet = bg_dataSet
    #     return
    def __init__(self,):

        self._dataSet = []
        self._bg_dataSet = []
        return

    def set_data_number(self,data,bg_data):
        # self._dataSet.setdefault(dataSet,[]).append(dataSet)
        # dataSet.setdefault(c.value,[]).append(d1.value)

        self._dataSet.append(data)
        self._bg_dataSet.append(bg_data)
        return

class FitsBook():
    def __init__(self,):
        self._fitDataDict = defaultdict(FitsDataClass)
        return
    def get_fits_class(self,key):
        
        return self._fitDataDict[key]

def Get_experiment_book(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]
    first_row = ws[1]
    configuration_index = None
    d1_index = None
    d0_index = None

    for _, v in enumerate(first_row):
        print(v.value)
        if v.value == "Configuration":
            configuration_index = getColumnCoordinate(v)
        elif v.value == "Scan Number":
            d1_index = getColumnCoordinate(v)
        elif v.value == "BG Number":
            d0_index = getColumnCoordinate(v)
    if not(configuration_index and d1_index and d0_index):
        sys.exit("scan number error")
    configurations = ws[configuration_index][1:]
    d1s = ws[d1_index][1:]
    d0s = ws[d0_index][1:]

    index = 0
    book = FitsBook()
    for (c,d1,d0)in zip(configurations,d1s,d0s):
        d = book.get_fits_class(c.value)
        d.set_data_number(d1.value,d0.value)

    return book


  